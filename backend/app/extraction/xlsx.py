"""Spreadsheet parsing.

One worksheet becomes one page, labelled with the sheet name so evidence can
cite "Sheet: Transactions" instead of a page number that does not exist. Cell
values are kept as strings in the order they appear; interpreting them is the
extractor's job, not the parser's.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.extraction.base import ParsedDocument, ParsedPage, ParsedTable, ParsingError

# Guards against a spreadsheet with a stray value in row 100000 producing a
# multi-megabyte text blob that would then be sent to a model.
MAX_ROWS_PER_SHEET = 5000
MAX_COLS_PER_SHEET = 64


def parse_xlsx(path: str) -> ParsedDocument:
    if Path(path).suffix.lower() == ".xls":
        raise ParsingError(
            "legacy .xls is not supported; save the workbook as .xlsx",
            code="UNSUPPORTED_FILE",
        )
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ParsingError(f"could not open the workbook: {exc}", code="CORRUPTED_FILE") from exc

    pages: list[ParsedPage] = []
    try:
        for index, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]
            rows: list[list[str]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_index > MAX_ROWS_PER_SHEET:
                    break
                cells = [
                    "" if value is None else str(value).strip()
                    for value in row[:MAX_COLS_PER_SHEET]
                ]
                if any(cells):
                    rows.append(cells)

            table = ParsedTable(page_number=index, rows=rows, name=sheet_name)
            text = f"# Sheet: {sheet_name}\n{table.to_text()}" if rows else ""
            pages.append(
                ParsedPage(
                    page_number=index,
                    text=text,
                    has_text_layer=bool(rows),
                    needs_ocr=False,
                    tables=[table] if rows else [],
                    label=f"Sheet: {sheet_name}",
                )
            )
    finally:
        workbook.close()

    if not pages:
        raise ParsingError("the workbook contains no sheets", code="CORRUPTED_FILE")

    readable = any(page.has_text_layer for page in pages)
    return ParsedDocument(
        pages=pages,
        page_count=len(pages),
        source_format="xlsx",
        is_readable=readable,
        quality_flags=[] if readable else ["UNREADABLE"],
        metadata={"sheets": workbook.sheetnames if hasattr(workbook, "sheetnames") else []},
    )
